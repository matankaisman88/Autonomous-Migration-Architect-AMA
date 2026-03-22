from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

from ama.discovery import split_qualified_name
from ama.sanitize import has_rtl_script

# Hebrew, Arabic scripts — hide from primary Markdown cells for LTR terminals
_RTL_SCRIPTS_RE = re.compile(r"[\u0590-\u05FF\u0600-\u06FF\u0700-\u074F]")


def is_ascii_identifier(s: str) -> bool:
    s = s.strip()
    return bool(s) and all(ord(c) < 128 for c in s)


def mask_non_ascii_runs(s: str) -> str:
    """
    Replace RTL script characters with a single token so terminals do not show reversed words.
    Keeps ASCII, digits, punctuation.
    """
    if not s:
        return ""
    out = _RTL_SCRIPTS_RE.sub("[HE]", s)
    out = re.sub(r"(\[HE\]\s*)+", "[HE] ", out)
    return out.strip()


def legacy_source_summary(source_columns: list[str]) -> str:
    """
    Count-based summary for confirmed rows — no raw Hebrew in the primary cell.
    """
    if not source_columns:
        return "0"
    n = len(source_columns)
    ascii_n = sum(1 for s in source_columns if is_ascii_identifier(s))
    rtl_n = n - ascii_n
    if rtl_n == 0:
        return f"{n} ASCII"
    if ascii_n == 0:
        return f"{n} non-ASCII (Hebrew/Arabic hidden)"
    return f"{n} total ({ascii_n} ASCII + {rtl_n} non-ASCII hidden)"


def ascii_legacy_names_only(source_columns: list[str]) -> str:
    """Comma-separated legacy names that are safe for LTR terminals."""
    safe = [s for s in source_columns if is_ascii_identifier(s)]
    return ", ".join(safe) if safe else "—"


def sanitize_citations_for_markdown(citations: list[str], *, max_len: int = 220) -> str:
    """Secondary evidence: Hebrew masked, suitable for a narrow Notes column."""
    blob = " | ".join(citations)
    masked = mask_non_ascii_runs(blob)
    if len(masked) > max_len:
        return masked[: max_len - 3] + "..."
    return masked


def _esc(s: str) -> str:
    return s.replace("|", "\\|").replace("\n", " ")


def _trunc_cell(s: str, max_len: int = 140) -> str:
    s = s.strip()
    if len(s) <= max_len:
        return s
    return s[: max_len - 3] + "..."


def _safe_table_segment(table_full_name: str) -> str:
    s = table_full_name.replace(".", "_")
    s = re.sub(r'[<>:"/\\|?*]', "_", s)
    return s.strip() or "report"


def default_report_filename(table_full_name: str, extension: str) -> str:
    """`ama_report_<table>_<timestamp>.<ext>` — extension should include a leading dot."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = _safe_table_segment(table_full_name)
    ext = extension if extension.startswith(".") else f".{extension}"
    return f"ama_report_{base}_{ts}{ext}"


def _looks_like_directory_path_spec(user_spec: str) -> bool:
    s = user_spec.rstrip()
    if not s:
        return False
    return s.endswith(("/", "\\"))


def resolve_report_output_path(
    user_spec: str,
    *,
    table_full_name: str,
    extension: str,
    cwd: Path | None = None,
) -> Path:
    """
    Resolve a user-provided output path.
    - ``user_spec == ""`` → default filename in ``cwd``.
    - Existing directory → default filename inside that directory.
    - Path ending with ``/`` or ``\\`` → treat as directory (created if missing).
    - File path without suffix → ``extension`` appended (e.g. ``.md`` / ``.json``).
    """
    cwd = cwd or Path.cwd()
    ext = extension if extension.startswith(".") else f".{extension}"
    default_name = default_report_filename(table_full_name, ext)

    if user_spec == "":
        return (cwd / default_name).resolve()

    raw = user_spec
    p = Path(raw).expanduser()
    if not p.is_absolute():
        p = (cwd / p).resolve()

    if p.exists() and p.is_dir():
        return (p / default_name).resolve()

    if _looks_like_directory_path_spec(raw):
        p = Path(str(raw).rstrip("/\\")).expanduser()
        if not p.is_absolute():
            p = (cwd / p).resolve()
        p.mkdir(parents=True, exist_ok=True)
        return (p / default_name).resolve()

    if not p.suffix:
        p = p.with_suffix(ext)

    return p.resolve()


def write_report_file(path: Path | str, content: str, *, encoding: str = "utf-8") -> Path:
    """Create parent directories and write UTF-8 text. Returns the resolved path."""
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(content, encoding=encoding)
    return out.resolve()


def format_cli_run_summary(
    payload: dict[str, Any],
    *,
    fmt: str,
    include_markdown_tip: bool = False,
) -> str:
    """Short terminal-friendly lines (no wide Markdown tables)."""
    lines: list[str] = []
    target = str(payload.get("target_table", ""))
    lines.append(f"Target table: {target}")
    lines.append(f"Queries matched: {payload.get('queries_matched', 0)}")
    lines.append(f"Format: {fmt}")
    md = payload.get("markdown_sections") or {}
    conf = md.get("confirmed") or []
    if conf:
        lines.append(f"Confirmed migrations (DDL): {len(conf)}")
    src = payload.get("column_name_source")
    if src:
        lines.append(f"Column name source: {src}")
    if fmt == "markdown" and include_markdown_tip:
        lines.append("")
        lines.append(
            "Tip: terminal line-wrapping can break Markdown tables. "
            "Use --out-file PATH or -o PATH to save the full report for viewing in an editor."
        )
    if fmt == "excel" and include_markdown_tip:
        lines.append("")
        lines.append(
            "Tip: Excel reports are not printed to the terminal. "
            "Use --format excel --out-file report.xlsx (or -o path ending in .xlsx) to export."
        )
    return "\n".join(lines)


def render_markdown_summary(payload: dict[str, Any]) -> str:
    """
    Canonical-first Markdown: confirmed rows keyed by Target DDL only; Hebrew only in masked notes.
    """
    lines: list[str] = []
    target = payload.get("target_table", "")
    lines.append("# AMA migration column report")
    lines.append("")
    lines.append(f"**Target table:** `{_esc(str(target))}`  ")
    lines.append(f"**Queries matched:** {payload.get('queries_matched', 0)}  ")
    src = payload.get("column_name_source", "")
    lines.append(f"**Column name source:** `{_esc(str(src))}`  ")
    lines.append("")
    lines.append("*Confirmed section shows **DDL names only** in the first column; legacy Hebrew is summarized, not spelled out.*")
    lines.append("")

    md = payload.get("markdown_sections") or {}
    conf = md.get("confirmed") or []
    rev = md.get("review") or []
    trash = md.get("trash") or []

    lines.append("## Confirmed migrations (DDL canonical)")
    lines.append("")
    lines.append("| Target DDL | # Sources | Confidence | Strategy | Source trace (ASCII only) | Notes / citations |")
    lines.append("| --- | ---: | --- | --- | --- | --- |")
    for row in conf:
        ddl = _esc(str(row.get("ddl", "")))
        nsrc = row.get("source_count", row.get("sources", ""))
        cf = row.get("confidence", "")
        st = _esc(str(row.get("strategy", "")))
        trace = _esc(str(row.get("source_trace", row.get("ascii_legacy_names", "—"))))
        notes = _trunc_cell(_esc(str(row.get("notes", row.get("citations_safe", "")))))
        lines.append(f"| `{ddl}` | {nsrc} | {cf} | {st} | {trace} | {notes} |")
    if not conf:
        lines.append("| — | — | — | — | — | — |")
    lines.append("")

    lines.append("## Human review required (medium confidence)")
    lines.append("")
    lines.append("| Legacy (masked) | Suggested DDL | Confidence | Note (masked) |")
    lines.append("| --- | --- | --- | --- |")
    for row in rev:
        leg = _esc(mask_non_ascii_runs(str(row.get("legacy", ""))))
        lines.append(
            f"| `{leg}` | `{_esc(str(row.get('suggested_ddl', '')))}` | "
            f"{row.get('confidence', '')} | {_esc(mask_non_ascii_runs(str(row.get('note', ''))))} |"
        )
    if not rev:
        lines.append("| — | — | — | — |")
    lines.append("")

    lines.append("## Discarded / unmapped (low signal or rejected merge)")
    lines.append("")
    lines.append("| Legacy (masked) | Best-guess DDL | Confidence | Reason (masked) |")
    lines.append("| --- | --- | --- | --- |")
    for row in trash:
        leg = _esc(mask_non_ascii_runs(str(row.get("legacy", ""))))
        lines.append(
            f"| `{leg}` | `{_esc(str(row.get('suggested_ddl', '')))}` | "
            f"{row.get('confidence', '')} | {_esc(mask_non_ascii_runs(str(row.get('note', ''))))} |"
        )
    if not trash:
        lines.append("| — | — | — | — |")
    lines.append("")

    imp = payload.get("importance_ddl") or payload.get("columns") or []
    if imp:
        lines.append("## Importance (DDL identifiers only)")
        lines.append("")
        lines.append("| DDL column | Importance score | Dead candidate |")
        lines.append("| --- | --- | --- |")
        for row in imp:
            if not isinstance(row, dict):
                continue
            col = str(row.get("column", ""))
            if not is_ascii_identifier(col):
                continue
            col_e = _esc(col)
            sc = row.get("importance_score", "")
            dc = row.get("dead_candidate", "")
            lines.append(f"| `{col_e}` | {sc} | {dc} |")
        lines.append("")

    um = payload.get("unmapped_importance") or []
    if um:
        lines.append("## Unmapped / non-DDL — importance (masked identifiers)")
        lines.append("")
        lines.append("| Legacy (masked) | Importance score | Category |")
        lines.append("| --- | --- | --- |")
        for row in um:
            if not isinstance(row, dict):
                continue
            col = mask_non_ascii_runs(str(row.get("column", "")))
            lines.append(
                f"| `{_esc(col)}` | {row.get('importance_score', '')} | "
                f"{_esc(str(row.get('category', '')))} |"
            )
        lines.append("")

    lines.append("---")
    lines.append("*Non-ASCII legacy tokens appear as `[HE]` in Markdown for LTR terminal safety. Full strings remain in JSON (`alias_merge`).*")
    lines.append("")
    return "\n".join(lines)


class ExcelReportGenerator:
    """
    Managerial .xlsx export: Dashboard, Migration Map (confirmed + review), Trash.
    Legacy Hebrew is shown unmasked in the Legacy Context column; Target DDL stays canonical.
    With discovery payload: Executive Summary (domain matrix + table fact sheets), Database Inventory
    (domains, hyperlinks), Global Migration Map (grouped rows), Trash & Orphans.
    """

    _HEADER_FILL = "FF4472C4"
    _HEADER_FONT = "FFFFFFFF"
    _GREEN = "FFC6EFCE"
    _YELLOW = "FFFFEB9C"
    _RED = "FFFFC7CE"

    def __init__(self, payload: dict[str, Any]) -> None:
        self._payload = payload

    def write(self, path: Path | str) -> Path:
        try:
            import pandas as pd
        except ImportError as e:  # pragma: no cover
            raise ImportError(
                "Excel export requires pandas and openpyxl. Install: pip install pandas openpyxl"
            ) from e

        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)

        disc = self._payload.get("discovery") or {}
        if isinstance(disc, dict) and disc.get("enabled"):
            return self._write_discovery_workbook(out)

        dash = self._dashboard_df()
        mig = self._migration_map_df()
        trash = self._trash_df()

        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            dash.to_excel(writer, sheet_name="Dashboard", index=False)
            mig.to_excel(writer, sheet_name="Migration Map", index=False)
            trash.to_excel(writer, sheet_name="Trash", index=False)

            wb = writer.book
            self._format_sheet(wb["Dashboard"], confidence_col=None, legacy_col=None)
            self._format_sheet(
                wb["Migration Map"],
                confidence_col="Confidence",
                legacy_col="Legacy Context",
            )
            self._format_sheet(
                wb["Trash"],
                confidence_col="Confidence",
                legacy_col="Legacy",
            )

        return out.resolve()

    def _write_discovery_workbook(self, out: Path) -> Path:
        import pandas as pd

        inv = self._inventory_df()
        glob_m = self._migration_map_df()
        trash = self._trash_df_discovery()
        disc = self._payload.get("discovery") or {}
        hier = bool(disc.get("enabled"))
        legacy_col = "Legacy Name" if hier else "Legacy Context"

        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            inv.to_excel(writer, sheet_name="Database Inventory", index=False)
            glob_m.to_excel(writer, sheet_name="Global Migration Map", index=False)
            trash.to_excel(writer, sheet_name="Trash & Orphans", index=False)

            wb = writer.book
            self._prepend_executive_summary(wb)
            self._format_sheet(wb["Database Inventory"], confidence_col=None, legacy_col=None)
            self._apply_rtl_inventory_columns(wb["Database Inventory"])
            self._apply_technical_debt_outline(wb["Database Inventory"])
            self._format_sheet(
                wb["Global Migration Map"],
                confidence_col="Confidence",
                legacy_col=legacy_col,
            )
            self._format_sheet(
                wb["Trash & Orphans"],
                confidence_col="Confidence",
                legacy_col="Legacy",
            )
            self._apply_global_map_outline(wb["Global Migration Map"])
            self._apply_inventory_hyperlinks(wb)
            self._format_executive_summary_sheet(wb)

        return out.resolve()

    def _default_database(self) -> str:
        d = self._payload.get("discovery") or {}
        return str(d.get("default_database") or "")

    def _location_prefix(self) -> tuple[str, str, str, str]:
        d = self._payload.get("discovery") or {}
        default_db = str(d.get("default_database") or "")
        tk = str(d.get("target_key") or "")
        if tk:
            db, s, t = split_qualified_name(tk)
            if not db and default_db:
                db = default_db
            return db, s, t, tk
        tgt = str(self._payload.get("target_table") or "")
        if "." in tgt:
            a, _, b = tgt.partition(".")
            db = default_db if default_db else ""
            return db, a, b, f"{a}.{b}"
        return default_db, "", tgt, tgt

    def _prepend_executive_summary(self, wb) -> None:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        es = (self._payload.get("discovery") or {}).get("executive_summary") or {}
        dm = es.get("domain_matrix") or []
        facts = es.get("table_fact_sheets") or []
        ws = wb.create_sheet("Executive Summary", 0)
        thin = Side(style="thin", color="FFCCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_font = Font(bold=True, size=14)
        hdr_font = Font(bold=True, color=self._HEADER_FONT)
        hdr_fill = PatternFill(
            start_color=self._HEADER_FILL,
            end_color=self._HEADER_FILL,
            fill_type="solid",
        )
        ws["A1"] = "Executive Summary — Migration Portfolio"
        ws.merge_cells("A1:F1")
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        r = 3
        ws.cell(row=r, column=1, value="Complexity vs. importance by business domain")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).font = Font(bold=True, size=11)
        r += 1
        headers = [
            "Business Domain",
            "Importance (0–100)",
            "Complexity (0–100)",
            "Tables",
            "Importance bar",
            "Why this domain matters for migration",
        ]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        hdr_row = r
        r += 1
        for row in dm:
            imp = float(row.get("business_importance", 0) or 0)
            cx = float(row.get("migration_complexity", 0) or 0)
            bar_n = max(1, min(10, int(round(imp / 10))))
            bar = "█" * bar_n + "░" * max(0, 10 - bar_n)
            ws.cell(row=r, column=1, value=row.get("business_domain", ""))
            ws.cell(row=r, column=2, value=round(imp, 1))
            ws.cell(row=r, column=3, value=round(cx, 1))
            ws.cell(row=r, column=4, value=row.get("table_count", ""))
            ws.cell(row=r, column=5, value=bar)
            ws.cell(row=r, column=6, value=row.get("narrative", ""))
            for c in range(1, 7):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        r += 1
        ws.cell(row=r, column=1, value="Table fact sheets — highest-activity tables")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).font = Font(bold=True, size=11)
        r += 1
        hf = ["Full Qualified Name", "Domain", "Query Count", "Business description"]
        for c, h in enumerate(hf, 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        r += 1
        for row in facts:
            ws.cell(row=r, column=1, value=row.get("full_qualified_name", ""))
            ws.cell(row=r, column=2, value=row.get("business_domain", ""))
            ws.cell(row=r, column=3, value=row.get("query_count", ""))
            ws.cell(row=r, column=4, value=row.get("business_description", ""))
            for c in range(1, 5):
                ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        risk_hot = es.get("risk_hotspots") or []
        if risk_hot:
            r += 1
            ws.cell(row=r, column=1, value="Lineage risk — cross-domain blast radius (additive)")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.cell(row=r, column=1).font = Font(bold=True, size=11)
            r += 1
            hrh = ["Table", "Blast radius score", "Domains touched", "Downstream reach"]
            for c, h in enumerate(hrh, 1):
                cell = ws.cell(row=r, column=c, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
            r += 1
            for row in risk_hot[:40]:
                if not isinstance(row, dict):
                    continue
                ws.cell(row=r, column=1, value=row.get("table", ""))
                ws.cell(row=r, column=2, value=row.get("blast_radius_score", ""))
                ws.cell(row=r, column=3, value=", ".join(row.get("domains_touched") or []))
                ws.cell(row=r, column=4, value=row.get("downstream_tables_reached", ""))
                r += 1
        ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1).coordinate
        for col_letter, w in (
            ("A", 26),
            ("B", 12),
            ("C", 12),
            ("D", 10),
            ("E", 16),
            ("F", 62),
        ):
            ws.column_dimensions[col_letter].width = w

    def _format_executive_summary_sheet(self, wb) -> None:
        if "Executive Summary" not in wb.sheetnames:
            return
        ws = wb["Executive Summary"]
        ws.sheet_view.showGridLines = True

    def _apply_technical_debt_outline(self, ws) -> None:
        hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        ps_col = hdr.get("Portfolio Section")
        if not ps_col or ws.max_row < 2:
            return
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=ps_col).value
            if str(v or "") == "Technical Debt":
                ws.row_dimensions[r].outline_level = 1

    def _inventory_df(self):
        import pandas as pd

        rows = (self._payload.get("discovery") or {}).get("inventory") or []
        df = pd.DataFrame(rows)
        if df.empty:
            return pd.DataFrame(
                columns=[
                    "Portfolio Section",
                    "Business Domain",
                    "Database",
                    "Schema",
                    "Table",
                    "Full Qualified Name",
                    "Business Description",
                    "Query Count",
                    "Column Count",
                    "Priority Score",
                    "Status",
                ]
            )
        df = df.rename(
            columns={
                "portfolio_section": "Portfolio Section",
                "business_domain": "Business Domain",
                "database": "Database",
                "schema": "Schema",
                "table": "Table",
                "full_name": "Full Qualified Name",
                "query_count": "Query Count",
                "column_count": "Column Count",
                "priority_score": "Priority Score",
                "business_description": "Business Description",
                "table_comment": "Metadata Comment",
            }
        )
        preferred = [
            "Portfolio Section",
            "Business Domain",
            "Database",
            "Schema",
            "Table",
            "Full Qualified Name",
            "Business Description",
            "Metadata Comment",
            "Query Count",
            "Column Count",
            "Priority Score",
            "Status",
        ]
        front = [c for c in preferred if c in df.columns]
        rest = [c for c in df.columns if c not in front]
        return df[front + rest]

    def _schema_breakdown_df(self):
        import pandas as pd

        sb = (self._payload.get("discovery") or {}).get("schema_breakdown") or []
        df = pd.DataFrame(sb)
        if df.empty:
            return df
        return df.rename(
            columns={
                "table_count": "Table count",
                "total_queries": "Total queries",
                "approx_pct_confirmed": "% Confirmed (target merge)",
                "approx_pct_review": "% Review (target merge)",
                "has_target_table": "Contains target table",
            }
        )

    def _trash_df_discovery(self):
        import pandas as pd

        am = self._payload.get("alias_merge") or {}
        rows: list[dict[str, Any]] = []
        for u in am.get("trash_candidates") or []:
            if not isinstance(u, dict):
                continue
            try:
                conf = float(u.get("merge_confidence", 0.0))
            except (TypeError, ValueError):
                conf = 0.0
            st = str(u.get("source_table", "") or "")
            db, s, t = split_qualified_name(st) if st else ("", "", "")
            if not db and self._default_database():
                db = self._default_database()
            rows.append(
                {
                    "Source Table": st,
                    "Database": db,
                    "Schema": s,
                    "Table": t,
                    "Legacy": str(u.get("legacy_name", "")),
                    "Suggested DDL": str(u.get("suggested_ddl", "")),
                    "Confidence": conf,
                    "Category": str(u.get("category", "trash")),
                    "Strategy": str(u.get("strategy", "")),
                    "Notes": str(u.get("citation", "")),
                }
            )
        rows.sort(key=lambda r: (str(r.get("Source Table", "")), -float(r.get("Confidence") or 0.0)))
        if not rows:
            return pd.DataFrame(
                columns=[
                    "Source Table",
                    "Database",
                    "Schema",
                    "Table",
                    "Legacy",
                    "Suggested DDL",
                    "Confidence",
                    "Category",
                    "Strategy",
                    "Notes",
                ]
            )
        return pd.DataFrame(rows)

    def _apply_rtl_inventory_columns(self, ws) -> None:
        from openpyxl.styles import Alignment

        hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        for name in ("Database", "Schema", "Table", "Full Qualified Name", "Business Domain", "Business Description"):
            ci = hdr.get(name)
            if not ci:
                continue
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=ci)
                v = cell.value
                if v is None:
                    continue
                s = str(v)
                if has_rtl_script(s):
                    cell.alignment = Alignment(wrap_text=True, vertical="top", readingOrder=2)
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _apply_global_map_outline(self, ws) -> None:
        hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        fn_col = hdr.get("Full Qualified Name")
        if not fn_col or ws.max_row < 2:
            return
        max_row = ws.max_row
        r = 2
        while r <= max_row:
            block_key = ws.cell(row=r, column=fn_col).value
            start = r
            r += 1
            while r <= max_row and ws.cell(row=r, column=fn_col).value == block_key:
                r += 1
            end = r - 1
            for rr in range(start + 1, end + 1):
                ws.row_dimensions[rr].outline_level = 1

    def _apply_inventory_hyperlinks(self, wb) -> None:
        from openpyxl.styles import Font

        if "Database Inventory" not in wb.sheetnames or "Global Migration Map" not in wb.sheetnames:
            return
        inv = wb["Database Inventory"]
        gmap = wb["Global Migration Map"]
        hdr_inv = {inv.cell(row=1, column=c).value: c for c in range(1, inv.max_column + 1)}
        hdr_g = {gmap.cell(row=1, column=c).value: c for c in range(1, gmap.max_column + 1)}
        fn_i = hdr_inv.get("Full Qualified Name")
        tbl_i = hdr_inv.get("Table")
        fn_g = hdr_g.get("Full Qualified Name")
        if not fn_i or not tbl_i or not fn_g:
            return
        first_row: dict[str, int] = {}
        for r in range(2, gmap.max_row + 1):
            k = str(gmap.cell(row=r, column=fn_g).value or "")
            if k and k not in first_row:
                first_row[k] = r
        for r in range(2, inv.max_row + 1):
            k = str(inv.cell(row=r, column=fn_i).value or "")
            if k in first_row:
                row = first_row[k]
                cell = inv.cell(row=r, column=tbl_i)
                cell.hyperlink = f"#'Global Migration Map'!A{row}"
                cell.font = Font(color="0563C1", underline="single")

    def _importance_map(self) -> dict[str, float]:
        out: dict[str, float] = {}
        for r in self._payload.get("importance_ddl") or []:
            if not isinstance(r, dict):
                continue
            c = str(r.get("column", ""))
            st = str(r.get("source_table", "") or "")
            if not c:
                continue
            try:
                v = float(r.get("importance_score", 0.0))
            except (TypeError, ValueError):
                v = 0.0
            out[c] = v
            if st:
                out[f"{st}::{c}"] = v
        return out

    def _dashboard_df(self):
        import pandas as pd

        p = self._payload
        target = str(p.get("target_table", ""))
        qm = int(p.get("queries_matched", 0) or 0)
        cols = p.get("columns") or []
        imp_ddl = p.get("importance_ddl") or []
        total_columns = len(imp_ddl) if imp_ddl else len(cols)

        am = p.get("alias_merge") or {}
        n_conf = len(am.get("merged_entities") or [])
        n_rev = len(am.get("review_candidates") or [])
        n_trash = len(am.get("trash_candidates") or [])

        dead = 0
        for r in imp_ddl:
            if isinstance(r, dict) and r.get("dead_candidate"):
                dead += 1

        tot_imp = 0.0
        for r in imp_ddl:
            if isinstance(r, dict):
                try:
                    tot_imp += float(r.get("importance_score", 0.0))
                except (TypeError, ValueError):
                    pass

        pct_conf = (100.0 * n_conf / total_columns) if total_columns else 0.0
        pct_dead = (100.0 * dead / total_columns) if total_columns else 0.0

        metrics = [
            ("Target table", target),
            ("Queries matched", qm),
            ("Total columns (scope)", total_columns),
            ("Confirmed migrations", n_conf),
            ("% Confirmed", round(pct_conf, 2)),
            ("Review (medium confidence)", n_rev),
            ("Trash / low-signal rows", n_trash),
            ("Dead candidates (importance)", dead),
            ("% Dead candidates", round(pct_dead, 2)),
            ("Total importance (sum of scores)", round(tot_imp, 4)),
        ]
        return pd.DataFrame(metrics, columns=["Metric", "Value"])

    def _migration_map_df(self):
        import pandas as pd

        imp = self._importance_map()
        am = self._payload.get("alias_merge") or {}
        disc = self._payload.get("discovery") or {}
        hier = bool(disc.get("enabled"))
        ddl_k = "Canonical DDL Name" if hier else "Target DDL"
        leg_k = "Legacy Name" if hier else "Legacy Context"
        rows: list[dict[str, Any]] = []

        def _loc_from_source(st: str) -> tuple[str, str, str, str]:
            if st:
                db, s, t = split_qualified_name(st)
                if not db and self._default_database():
                    db = self._default_database()
                return db, s, t, st
            return self._location_prefix()

        def _imp_lookup(st: str, ddl: str) -> float | str:
            if st and hier:
                return imp.get(f"{st}::{ddl}", imp.get(ddl, ""))
            return imp.get(ddl, "")

        for ent in am.get("merged_entities") or []:
            if not isinstance(ent, dict):
                continue
            ddl = str(ent.get("canonical_column", ""))
            st = str(ent.get("source_table", "") or "")
            db, sc, tb, fq = _loc_from_source(st)
            sources = ent.get("source_columns") or []
            legacy_ctx = ", ".join(str(s) for s in sources) if sources else ""
            strat = ent.get("strategies") or []
            strat_s = ",".join(str(x) for x in strat) if isinstance(strat, list) else str(strat)
            cites = ent.get("citations") or []
            notes = " | ".join(str(c) for c in cites) if isinstance(cites, list) else str(cites)
            try:
                conf = float(ent.get("merge_confidence", 0.0))
            except (TypeError, ValueError):
                conf = 0.0
            base = {
                "Section": "Confirmed",
                ddl_k: ddl,
                leg_k: legacy_ctx,
                "Confidence": conf,
                "Importance": _imp_lookup(st, ddl),
                "Strategy": strat_s,
                "Notes": notes,
            }
            if hier:
                base = {
                    "Database": db,
                    "Schema": sc,
                    "Table": tb,
                    "Full Qualified Name": fq,
                    **base,
                }
            rows.append(base)

        for u in am.get("review_candidates") or []:
            if not isinstance(u, dict):
                continue
            ddl = str(u.get("suggested_ddl", ""))
            leg = str(u.get("legacy_name", ""))
            st = str(u.get("source_table", "") or "")
            db, sc, tb, fq = _loc_from_source(st)
            try:
                conf = float(u.get("merge_confidence", 0.0))
            except (TypeError, ValueError):
                conf = 0.0
            base = {
                "Section": "Review",
                ddl_k: ddl,
                leg_k: leg,
                "Confidence": conf,
                "Importance": _imp_lookup(st, ddl),
                "Strategy": str(u.get("strategy", "")),
                "Notes": str(u.get("citation", "")),
            }
            if hier:
                base = {
                    "Database": db,
                    "Schema": sc,
                    "Table": tb,
                    "Full Qualified Name": fq,
                    **base,
                }
            rows.append(base)

        def _sort_key(r: dict[str, Any]) -> tuple[Any, ...]:
            sec = r.get("Section", "")
            if hier:
                fq = str(r.get("Full Qualified Name", ""))
                if sec == "Confirmed":
                    impv = float(r.get("Importance") or 0.0)
                    return (fq, 0, -impv, str(r.get(ddl_k, "")))
                cf = float(r.get("Confidence") or 0.0)
                return (fq, 1, -cf, str(r.get(leg_k, "")))
            if sec == "Confirmed":
                impv = float(r.get("Importance") or 0.0)
                return (0, -impv, str(r.get(ddl_k, "")))
            cf = float(r.get("Confidence") or 0.0)
            return (1, -cf, str(r.get(leg_k, "")))

        rows.sort(key=_sort_key)
        if not rows:
            cols = [
                "Section",
                ddl_k,
                leg_k,
                "Confidence",
                "Importance",
                "Strategy",
                "Notes",
            ]
            if hier:
                cols = ["Database", "Schema", "Table", "Full Qualified Name"] + cols
            return pd.DataFrame(columns=cols)
        return pd.DataFrame(rows)

    def _trash_df(self):
        import pandas as pd

        am = self._payload.get("alias_merge") or {}
        rows: list[dict[str, Any]] = []
        for u in am.get("trash_candidates") or []:
            if not isinstance(u, dict):
                continue
            try:
                conf = float(u.get("merge_confidence", 0.0))
            except (TypeError, ValueError):
                conf = 0.0
            rows.append(
                {
                    "Legacy": str(u.get("legacy_name", "")),
                    "Suggested DDL": str(u.get("suggested_ddl", "")),
                    "Confidence": conf,
                    "Category": str(u.get("category", "trash")),
                    "Strategy": str(u.get("strategy", "")),
                    "Notes": str(u.get("citation", "")),
                }
            )
        rows.sort(key=lambda r: -float(r.get("Confidence") or 0.0))
        if not rows:
            return pd.DataFrame(
                columns=["Legacy", "Suggested DDL", "Confidence", "Category", "Strategy", "Notes"]
            )
        return pd.DataFrame(rows)

    def _format_sheet(
        self,
        ws,
        *,
        confidence_col: str | None,
        legacy_col: str | None,
    ) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill(
            start_color=self._HEADER_FILL,
            end_color=self._HEADER_FILL,
            fill_type="solid",
        )
        header_font = Font(bold=True, color=self._HEADER_FONT)

        max_col = ws.max_column
        max_row = ws.max_row
        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "A2"

        hdr = {ws.cell(row=1, column=c).value: c for c in range(1, max_col + 1)}

        if confidence_col and confidence_col in hdr and max_row >= 2:
            from openpyxl.formatting.rule import FormulaRule

            ci = hdr[confidence_col]
            col_letter = get_column_letter(ci)
            green = PatternFill(start_color=self._GREEN, end_color=self._GREEN, fill_type="solid")
            yellow = PatternFill(start_color=self._YELLOW, end_color=self._YELLOW, fill_type="solid")
            red = PatternFill(start_color=self._RED, end_color=self._RED, fill_type="solid")
            r0 = 2
            rng = f"{col_letter}{r0}:{col_letter}{max_row}"
            c = col_letter
            # Conditional formatting (editable in Excel): red < 0.4, yellow 0.4–0.8, green >= 0.8
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=[f"{c}{r0}<0.4"], fill=red),
            )
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=[f"AND({c}{r0}>=0.4,{c}{r0}<0.8)"], fill=yellow),
            )
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=[f"{c}{r0}>=0.8"], fill=green),
            )

        if legacy_col and legacy_col in hdr and max_row >= 2:
            li = hdr[legacy_col]
            rtl_align = Alignment(wrap_text=True, vertical="top", readingOrder=2)
            ltr_align = Alignment(wrap_text=True, vertical="top")
            for r in range(2, max_row + 1):
                cell = ws.cell(row=r, column=li)
                val = cell.value
                if val is None:
                    continue
                s = str(val)
                cell.alignment = rtl_align if has_rtl_script(s) else ltr_align

        for c in range(1, max_col + 1):
            letter = get_column_letter(c)
            maxlen = 0
            for r in range(1, max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    maxlen = max(maxlen, len(str(v)))
            ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), 55)


def write_excel_report(payload: dict[str, Any], path: Path | str) -> Path:
    """Write managerial Excel workbook; returns resolved path."""
    gen = ExcelReportGenerator(payload)
    return gen.write(path)
