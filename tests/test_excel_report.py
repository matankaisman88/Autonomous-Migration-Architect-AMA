from __future__ import annotations

from pathlib import Path

import openpyxl

from ama.reports import write_excel_report


def test_write_excel_report_multi_sheet(tmp_path: Path) -> None:
    payload = {
        "migration_context": "sales.orders",
        "queries_matched": 10,
        "column_name_source": "ddl",
        "importance_ddl": [
            {"column": "status", "importance_score": 1.0, "dead_candidate": False},
        ],
        "columns": [],
        "alias_merge": {
            "merged_entities": [
                {
                    "canonical_column": "status",
                    "source_columns": ["status", "סטטוס"],
                    "merge_confidence": 0.98,
                    "strategies": ["glossary", "exact_ddl"],
                    "citations": ["glossary hit"],
                    "hitl": False,
                }
            ],
            "review_candidates": [
                {
                    "legacy_name": "maybe_col",
                    "suggested_ddl": "amount",
                    "merge_confidence": 0.55,
                    "category": "review",
                    "citation": "weak",
                    "strategy": "vector",
                }
            ],
            "trash_candidates": [
                {
                    "legacy_name": "junk",
                    "suggested_ddl": "order_id",
                    "merge_confidence": 0.2,
                    "category": "trash",
                    "citation": "low",
                    "strategy": "lexical",
                }
            ],
        },
    }
    path = tmp_path / "out.xlsx"
    out = write_excel_report(payload, path)
    assert out == path.resolve()

    wb = openpyxl.load_workbook(path)
    assert "Dashboard" in wb.sheetnames
    assert "Migration Map" in wb.sheetnames
    assert "Trash" in wb.sheetnames
    mm = wb["Migration Map"]
    assert mm["C2"].value == "status, סטטוס"
    assert mm.freeze_panes == "A2"


def test_write_excel_report_discovery_executive_summary_first_sheet(tmp_path: Path) -> None:
    payload = {
        "migration_context": "sales.orders",
        "queries_matched": 3,
        "column_name_source": "ddl",
        "alias_merge": {"merged_entities": [], "review_candidates": [], "trash_candidates": []},
        "discovery": {
            "enabled": True,
            "default_database": "CORP_CATALOG",
            "inventory": [
                {
                    "database": "CORP_CATALOG",
                    "schema": "PROD_SALES",
                    "table": "Orders",
                    "full_name": "PROD_SALES.Orders",
                    "query_count": 10,
                    "column_count": 2,
                    "priority_score": 100.0,
                    "status": "Ready for Migration",
                    "portfolio_section": "Core Business",
                    "business_domain": "Finance",
                    "business_description": "First sentence about Orders. Second sentence about impact.",
                }
            ],
            "executive_summary": {
                "domain_matrix": [
                    {
                        "business_domain": "Finance",
                        "business_importance": 88.0,
                        "migration_complexity": 42.0,
                        "table_count": 1,
                        "narrative": "Finance domain drives revenue cutover.",
                    }
                ],
                "table_fact_sheets": [
                    {
                        "full_qualified_name": "PROD_SALES.Orders",
                        "business_domain": "Finance",
                        "query_count": 10,
                        "business_description": "Fact sheet line one. Fact sheet line two.",
                    }
                ],
            },
        },
    }
    path = tmp_path / "disc.xlsx"
    out = write_excel_report(payload, path)
    assert out == path.resolve()
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames[0] == "Executive Summary"
    assert "Database Inventory" in wb.sheetnames
    inv = wb["Database Inventory"]
    assert inv["A1"].value == "Portfolio Section"
    assert inv["B1"].value == "Business Domain"
    assert inv["C1"].value == "Database"
    assert inv["C2"].value == "CORP_CATALOG"
